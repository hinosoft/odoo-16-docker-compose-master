# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""

from django import template
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
from openpyxl import Workbook
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

@login_required(login_url="/login/")
def index(request):
    context = {'segment': 'index'}

    html_template = loader.get_template('home/index.html')
    return HttpResponse(html_template.render(context, request))


@login_required(login_url="/login/")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))
import openpyxl
def download_report(request):
    
    wb = Workbook()
    wb = openpyxl.load_workbook('apps/home/Summary time attendant.xlsx')
    ws = wb.create_sheet()
    ws.title = "Pi"
    for start_row in range(1,100):
        ws.cell(row=start_row, column=6).value = start_row
    response = HttpResponse(content_type='application/vnd.ms-excel')
    wb.save(response)
    return response

